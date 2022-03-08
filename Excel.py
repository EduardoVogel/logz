from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.styles import PatternFill
import os
from datetime import datetime

caminho = "req_v4.xlsx"
data = datetime.today().strftime('%d/%m/%Y')

fill = PatternFill(patternType="solid", start_color="AEAAAA") # Preenchimento de cor pura

if os.path.isfile(caminho) and os.access(caminho, os.R_OK):
    print("Foi possivel acessar o arquivo, utilizando dados arquivo PC")
    wb = load_workbook(caminho)
else:
    print("Não foi possivel acessar o arquivo, utilizando arquivo próprio")
    wb = load_workbook('Backup.xlsx')

# Seleciona a active Sheet
#print(wb.sheetnames)

ws6 = wb["JAN"]
ws6["B23"].value = data+'ifjeifje'
ws6['A23'].fill = fill

# Salva arquivo (Se não colocar o caminho complete, ele salva
wb.save('req_v4.xlsx')